from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
import math
import re
from statistics import median
from typing import Any, Iterable

from openpyxl import load_workbook

MAX_TABLE_ROWS = 50_000
MAX_PREVIEW_ROWS = 8
MAX_DISTINCT_VALUES = 40

CATEGORY_HINTS = {
    "size", "message", "msg", "packet", "frame", "framesize", "input", "test",
    "operation", "collective", "workload", "configuration", "config", "mode",
    "batch", "model", "sequence", "date", "time", "name", "type", "category",
}
DIMENSION_HINTS = {
    "context", "contexts", "thread", "threads", "trial", "iteration", "run", "rank",
    "qp", "queue pair", "node", "nodes", "gpu", "gpus", "soc", "datatype", "data type",
    "build", "platform", "nic", "firmware", "driver", "gbs", "seq len", "seq_len",
    "gradacc", "tp", "pp", "ep", "world size", "batch size", "timed msgs", "timed messages",
}
IDENTIFIER_HINTS = {
    "trial", "iteration", "run", "rank", "id", "index", "timed msgs", "timed messages",
}
METRIC_HINTS = {
    "bandwidth", "throughput", "latency", "rate", "mfu", "utilization", "efficiency",
    "performance", "score", "index", "time to solution", "time_to_solution", "duration",
    "loss", "errors", "count", "fps", "gbps", "gb/s", "msg/s", "messages/s", "value",
}
DELTA_HINTS = {
    "diff", "difference", "delta", "change", "improvement", "regression", "gain", "uplift",
    "vs baseline", "%", "percent",
}
ANNOTATION_HINTS = {
    "note", "notes", "comment", "comments", "status", "image", "url", "link", "description",
    "commit", "sha", "owner", "author",
}
GROUP_HINTS = {
    "platform", "build", "soc", "model", "nic", "driver", "firmware", "context", "thread",
    "qp", "node", "gpu", "datatype", "mode", "config", "gbs", "batch",
}

SIZE_VALUE_RE = re.compile(
    r"^\s*[+-]?\d+(?:\.\d+)?\s*(?:[kmgtp](?:i?b)?|bytes?|b)\s*$", re.IGNORECASE
)
PERFORMANCE_NUMBER_RE = re.compile(
    r"([+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)\s*"
    r"(gb/s|gbps|mb/s|mbps|kb/s|kbps|ns|us|µs|ms|fps|msg/s|messages/s|%)",
    re.IGNORECASE,
)
NUMBER_WITH_UNIT_RE = re.compile(
    r"^\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?)\s*"
    r"(%|(?:[kmgtp]?(?:i?b|b))(?:/s|ps)?|gb/s|gbps|mb/s|mbps|kb/s|kbps|"
    r"ns|us|µs|ms|s|fps|msg/s|messages/s)?(?:\s*\([^)]*\))?\s*$",
    re.IGNORECASE,
)


def _normalized_header(value: str) -> str:
    text = re.sub(r"[_\-]+", " ", value.lower())
    text = re.sub(r"\([^)]*\)", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _contains_hint(text: str, hints: set[str]) -> bool:
    normalized = _normalized_header(text)
    return any(hint in normalized for hint in hints)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _coerce_number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            return None
        return numeric if math.isfinite(numeric) else None
    text = str(value).replace(",", "").strip()
    if not text:
        return None
    try:
        numeric = float(text)
        return numeric if math.isfinite(numeric) else None
    except ValueError:
        pass
    match = NUMBER_WITH_UNIT_RE.match(text)
    if not match:
        performance_matches = PERFORMANCE_NUMBER_RE.findall(text)
        if performance_matches:
            try:
                numeric = float(performance_matches[-1][0])
            except ValueError:
                return None
            return numeric if math.isfinite(numeric) else None
        return None
    try:
        numeric = float(match.group(1))
    except ValueError:
        return None
    return numeric if math.isfinite(numeric) else None


def _is_number(value: Any) -> bool:
    return _coerce_number(value) is not None


def _col_letter(index_1_based: int) -> str:
    result = ""
    value = index_1_based
    while value:
        value, rem = divmod(value - 1, 26)
        result = chr(65 + rem) + result
    return result


def _split_segments(columns: list[int], max_gap: int = 1) -> list[tuple[int, int]]:
    if not columns:
        return []
    segments: list[tuple[int, int]] = []
    start = previous = columns[0]
    for current in columns[1:]:
        if current - previous <= max_gap:
            previous = current
            continue
        if previous - start + 1 >= 2:
            segments.append((start, previous))
        start = previous = current
    if previous - start + 1 >= 2:
        segments.append((start, previous))
    return segments


def _unique_headers(raw_headers: Iterable[Any]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for position, value in enumerate(raw_headers, start=1):
        base = _clean_text(value) or f"Column {position}"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base} ({counts[base]})")
    return result


def _meaningful_title(value: Any) -> str:
    text = _clean_text(value)
    if not text or len(text) > 160:
        return ""
    lower = text.lower()
    if lower.startswith(("http://", "https://", "image:", "image ", "#")):
        return ""
    if _coerce_number(text) is not None:
        return ""
    if len(re.findall(r"[A-Za-z]", text)) < 2:
        return ""
    return text


def _nearest_title(rows: list[list[Any]], header_row_index: int, start_col: int, end_col: int) -> str:
    for row_index in range(header_row_index - 1, max(-1, header_row_index - 6), -1):
        values = [
            _meaningful_title(rows[row_index][column])
            for column in range(max(0, start_col - 1), min(end_col + 2, len(rows[row_index])))
        ]
        values = [value for value in values if value]
        if not values:
            continue
        short_values = [value for value in values if len(value) <= 90]
        if not short_values:
            continue
        title = " · ".join(short_values[:2])
        if len(title) <= 150:
            return title
    return ""


def _kind_for(sheet_name: str, title: str, headers: list[str]) -> str:
    text = " ".join([sheet_name, title, *headers]).lower()
    if any(token in text for token in ("all-reduce", "allreduce", "all2all", "alltoall", "rccl", "reducescatter", "reduce scatter", "allgather")):
        return "RCCL"
    if any(token in text for token in ("ib write", "ib read", "infiniband", "ib_", "queue pair", " qp", "qp ")):
        return "IB"
    if any(token in text for token in ("rocshmem", "deep ep", "deepep", "mori", "nonblocking put", "blocking put", "nonblocking get", "wave put", "gda")):
        return "rocSHMEM"
    if any(token in text for token in ("packet sweep", "pkt sweep", "framesize", "frame size", "line rate", "frame loss")):
        return "Packet Sweep"
    if any(token in text for token in ("time_to_solution", "time to solution", "mfu", "training", "workload", "gradient sync", "sparsity", "seq_len")):
        return "Training"
    if any(token in text for token in ("sweep", "message size", "input size")):
        return "Sweep"
    return "Benchmark"


def _looks_like_size(values: list[Any], header: str) -> bool:
    if _contains_hint(header, {"size", "framesize", "message", "packet", "input"}):
        return True
    populated = [_clean_text(value) for value in values if not _is_blank(value)]
    if not populated:
        return False
    matches = sum(1 for value in populated if SIZE_VALUE_RE.match(value))
    return matches / len(populated) >= 0.7


def _monotonic_ratio(values: list[float]) -> float:
    if len(values) < 3:
        return 0.0
    increasing = sum(1 for left, right in zip(values, values[1:]) if right >= left)
    decreasing = sum(1 for left, right in zip(values, values[1:]) if right <= left)
    return max(increasing, decreasing) / (len(values) - 1)


def _display_distinct(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return _clean_text(value)


@dataclass
class ColumnProfile:
    index: int
    header: str
    role: str
    data_type: str
    numeric_ratio: float
    unique_count: int
    populated_count: int
    distinct_values: list[str]
    distinct_values_truncated: bool
    metric_score: float
    dimension_score: float
    filterable: bool
    identifier: bool

    def public_dict(self) -> dict[str, Any]:
        return {
            "index": self.index,
            "header": self.header,
            "role": self.role,
            "data_type": self.data_type,
            "numeric_ratio": round(self.numeric_ratio, 3),
            "unique_count": self.unique_count,
            "populated_count": self.populated_count,
            "distinct_values": self.distinct_values,
            "distinct_values_truncated": self.distinct_values_truncated,
            "metric_score": round(self.metric_score, 2),
            "dimension_score": round(self.dimension_score, 2),
            "filterable": self.filterable,
            "identifier": self.identifier,
            "unit_family": _metric_unit_key(self.header),
        }


def _profile_column(index: int, header: str, values: list[Any]) -> ColumnProfile:
    populated = [value for value in values if not _is_blank(value)]
    numeric_values = [_coerce_number(value) for value in populated]
    numeric_values = [value for value in numeric_values if value is not None]
    numeric_ratio = len(numeric_values) / max(1, len(populated))
    distinct_seen: dict[str, None] = {}
    for value in populated:
        text = _display_distinct(value)
        if text not in distinct_seen:
            distinct_seen[text] = None
    all_distinct = list(distinct_seen)
    unique_count = len(all_distinct)
    distinct_values = all_distinct[:MAX_DISTINCT_VALUES]
    unique_ratio = unique_count / max(1, len(populated))
    header_norm = _normalized_header(header)
    is_size = _looks_like_size(populated, header)
    is_percent = "%" in header or "percent" in header_norm or (
        populated and sum(1 for value in populated if isinstance(value, str) and "%" in value) / len(populated) >= 0.5
    )
    date_count = sum(1 for value in populated if isinstance(value, (date, datetime)))
    if is_size:
        data_type = "size"
    elif is_percent:
        data_type = "percent"
    elif date_count / max(1, len(populated)) >= 0.7:
        data_type = "date"
    elif numeric_ratio >= 0.85:
        data_type = "number"
    elif numeric_ratio >= 0.25:
        data_type = "mixed"
    else:
        data_type = "text"

    annotation = _contains_hint(header, ANNOTATION_HINTS)
    delta = _contains_hint(header, DELTA_HINTS)
    identifier = header_norm in IDENTIFIER_HINTS or any(
        header_norm.startswith(f"{hint} ") for hint in IDENTIFIER_HINTS
    )
    dimension_hint = _contains_hint(header, DIMENSION_HINTS | CATEGORY_HINTS)
    metric_hint = _contains_hint(header, METRIC_HINTS)
    repeated = unique_ratio <= 0.65

    if annotation:
        role = "annotation"
    elif delta and numeric_ratio >= 0.35:
        role = "delta"
    elif is_size or data_type in {"text", "date"}:
        role = "dimension"
    elif identifier:
        role = "dimension"
    elif dimension_hint and repeated:
        role = "dimension"
    elif numeric_ratio >= 0.55:
        role = "metric"
    else:
        role = "dimension"

    metric_score = numeric_ratio * 5.0
    if metric_hint:
        metric_score += 5.5
    if any(token in header_norm for token in ("bandwidth", "throughput", "performance", "score", "index")):
        metric_score += 2.0
    raw_header = header.lower()
    if "gbps" in raw_header:
        metric_score += 4.5
    elif "gb/s" in raw_header:
        metric_score += 4.0
    elif any(token in raw_header for token in ("mbps", "mb/s", "kbps", "kb/s")):
        metric_score += 3.0
    if "mfu" in header_norm:
        metric_score += 4.0
    if "fps" in raw_header or "msg/s" in raw_header or "messages/s" in raw_header:
        metric_score += 1.5
    if any(token in header_norm for token in ("latency", "time to solution", "time_to_solution")):
        metric_score += 1.5
    if any(token in header_norm for token in ("loss", "errors", "count")) and not any(
        token in header_norm for token in ("throughput", "bandwidth")
    ):
        metric_score -= 1.5
    if delta:
        metric_score -= 3.0
    if identifier or (dimension_hint and repeated):
        metric_score -= 5.0
    if annotation:
        metric_score -= 20.0

    dimension_score = 0.0
    if _contains_hint(header, CATEGORY_HINTS):
        dimension_score += 8.0
    if dimension_hint:
        dimension_score += 4.0
    if is_size:
        dimension_score += 8.0
    if data_type in {"text", "date"}:
        dimension_score += 4.0
    if repeated:
        dimension_score += 2.5
    if numeric_values:
        dimension_score += _monotonic_ratio(numeric_values) * 1.5
    if metric_hint and not is_size:
        dimension_score -= 5.0
    if annotation:
        dimension_score -= 20.0

    filterable = role == "dimension" and 1 < unique_count <= MAX_DISTINCT_VALUES
    return ColumnProfile(
        index=index,
        header=header,
        role=role,
        data_type=data_type,
        numeric_ratio=numeric_ratio,
        unique_count=unique_count,
        populated_count=len(populated),
        distinct_values=distinct_values,
        distinct_values_truncated=unique_count > len(distinct_values),
        metric_score=metric_score,
        dimension_score=dimension_score,
        filterable=filterable,
        identifier=identifier,
    )


def _metric_unit_key(header: str) -> str:
    """Return a conservative unit/family key so defaults do not mix unlike scales."""
    text = header.lower().replace("µ", "u")
    normalized = _normalized_header(header)
    if "gbps" in text:
        return "gbps"
    if "gb/s" in text:
        return "gb_per_s"
    if "mbps" in text:
        return "mbps"
    if "mb/s" in text:
        return "mb_per_s"
    if "kbps" in text:
        return "kbps"
    if "kb/s" in text:
        return "kb_per_s"
    if "fps" in text:
        return "fps"
    if "msg/s" in text or "messages/s" in text:
        return "messages_per_s"
    if "mfu" in normalized or "efficiency" in normalized or "utilization" in normalized:
        return "percent"
    if "%" in text or "percent" in normalized or "line rate" in normalized:
        return "percent"
    if "latency" in normalized:
        for unit in ("ns", "us", "ms"):
            if re.search(rf"(?:^|[^a-z]){unit}(?:[^a-z]|$)", text):
                return f"latency_{unit}"
        return "latency"
    if "time_to_solution" in text or "time to solution" in normalized:
        return "duration_days" if "day" in text else "duration"
    if "duration" in normalized:
        return "duration"
    if "bandwidth" in normalized:
        return "bandwidth"
    if "throughput" in normalized:
        return "throughput"
    if any(token in normalized for token in ("loss", "error", "count")):
        return "count"
    return "generic"


def _suggest_y_axis(headers: list[str]) -> str:
    text = " ".join(headers).lower().replace("µ", "u")
    normalized = _normalized_header(" ".join(headers))
    mappings = (
        ("Throughput (Gbps)", "throughput" in normalized and "gbps" in text),
        ("Bandwidth (Gbps)", "bandwidth" in normalized and "gbps" in text),
        ("Throughput (GB/s)", "throughput" in normalized and "gb/s" in text),
        ("Bandwidth (GB/s)", "bandwidth" in normalized and "gb/s" in text),
        ("Throughput (fps)", "fps" in text),
        ("Message rate (msg/s)", "msg/s" in text or "messages/s" in text),
        ("Latency (ns)", "latency" in normalized and re.search(r"(?:^|[^a-z])ns(?:[^a-z]|$)", text) is not None),
        ("Latency (us)", "latency" in normalized and re.search(r"(?:^|[^a-z])us(?:[^a-z]|$)", text) is not None),
        ("Latency (ms)", "latency" in normalized and re.search(r"(?:^|[^a-z])ms(?:[^a-z]|$)", text) is not None),
        ("MFU (%)", "mfu" in normalized),
        ("Time to solution (days)", ("time_to_solution" in text or "time to solution" in normalized) and "day" in text),
        ("Time to solution", "time_to_solution" in text or "time to solution" in normalized),
        ("Frame loss (%)", "loss" in normalized and ("%" in text or "percent" in normalized)),
        ("Line rate (%)", "line rate" in normalized and ("%" in text or "percent" in normalized)),
        ("Bandwidth", "bandwidth" in normalized),
        ("Throughput", "throughput" in normalized),
        ("Latency", "latency" in normalized),
        ("Performance Index", "performance index" in normalized or normalized == "index"),
    )
    for label, matches in mappings:
        if matches:
            return label
    return "Value"


def _category_unique_count(profile: ColumnProfile) -> int:
    return profile.unique_count


@dataclass
class DetectedTable:
    table_id: str
    sheet_name: str
    sheet_state: str
    kind: str
    title: str
    display_name: str
    header_row: int
    data_start_row: int
    data_end_row: int
    start_col: int
    end_col: int
    headers: list[str]
    rows: list[list[Any]]
    column_profiles: list[ColumnProfile]
    suggested_series_indexes: list[int]
    suggested_category_index: int
    suggested_group_by_indexes: list[int]
    suggested_chart_type: str
    suggested_aggregation: str
    suggested_show_data_table: bool
    suggested_x_axis_title: str
    suggested_y_axis_title: str
    confidence: int
    warnings: list[str]

    @property
    def range(self) -> str:
        return f"{_col_letter(self.start_col)}{self.header_row}:{_col_letter(self.end_col)}{self.data_end_row}"

    @property
    def recommended(self) -> bool:
        return self.confidence >= 58

    def public_dict(self) -> dict[str, Any]:
        return {
            "id": self.table_id,
            "sheet_name": self.sheet_name,
            "sheet_state": self.sheet_state,
            "kind": self.kind,
            "title": self.title,
            "display_name": self.display_name,
            "header_row": self.header_row,
            "data_start_row": self.data_start_row,
            "data_end_row": self.data_end_row,
            "range": self.range,
            "row_count": len(self.rows),
            "column_count": len(self.headers),
            "headers": self.headers,
            "column_profiles": [profile.public_dict() for profile in self.column_profiles],
            "suggested_series_indexes": self.suggested_series_indexes,
            "suggested_category_index": self.suggested_category_index,
            "suggested_group_by_indexes": self.suggested_group_by_indexes,
            "suggested_chart_type": self.suggested_chart_type,
            "suggested_aggregation": self.suggested_aggregation,
            "suggested_show_data_table": self.suggested_show_data_table,
            "suggested_x_axis_title": self.suggested_x_axis_title,
            "suggested_y_axis_title": self.suggested_y_axis_title,
            "confidence": self.confidence,
            "recommended": self.recommended,
            "warnings": self.warnings,
            "preview_rows": [[_json_value(value) for value in row] for row in self.rows[:MAX_PREVIEW_ROWS]],
        }


@dataclass
class _Candidate:
    table: DetectedTable
    r0: int
    r1: int
    c0: int
    c1: int


def _collect_rows(
    normalized: list[list[Any]], header_index: int, start_col: int, end_col: int
) -> tuple[list[list[Any]], list[int], bool]:
    candidate_rows: list[list[Any]] = []
    row_numbers: list[int] = []
    truncated = False
    blank_run = 0
    weak_run = 0
    preamble_rows = 0
    for data_index in range(header_index + 1, len(normalized)):
        slice_values = normalized[data_index][start_col : end_col + 1]
        populated = sum(1 for value in slice_values if not _is_blank(value))
        numeric = sum(1 for value in slice_values if _is_number(value))
        text_count = populated - numeric
        if not candidate_rows:
            if populated >= 2 and numeric >= 1:
                candidate_rows.append(slice_values)
                row_numbers.append(data_index)
                continue
            preamble_rows += 1
            if preamble_rows > 3:
                break
            continue

        if populated == 0:
            blank_run += 1
            if blank_run >= 2:
                break
            continue
        blank_run = 0

        strong_new_header = numeric == 0 and text_count >= 2
        if strong_new_header:
            weak_run += 1
            if weak_run >= 1:
                break
            continue
        weak_run = 0
        if numeric >= 1 and populated >= 2:
            candidate_rows.append(slice_values)
            row_numbers.append(data_index)
            if len(candidate_rows) >= MAX_TABLE_ROWS:
                truncated = True
                break
        elif len(candidate_rows) >= 2:
            break
    return candidate_rows, row_numbers, truncated


def _trim_empty_edges(
    header_values: list[Any], rows: list[list[Any]], start_col: int
) -> tuple[list[Any], list[list[Any]], int, list[str]]:
    left = 0
    right = len(header_values) - 1
    removed_title_hints: list[str] = []
    while left <= right and all(_is_blank(row[left]) for row in rows):
        hint = _meaningful_title(header_values[left])
        if hint:
            removed_title_hints.append(hint)
        left += 1
    while right >= left and all(_is_blank(row[right]) for row in rows):
        hint = _meaningful_title(header_values[right])
        if hint:
            removed_title_hints.append(hint)
        right -= 1
    if left > right:
        return [], [], start_col, removed_title_hints
    return (
        header_values[left : right + 1],
        [row[left : right + 1] for row in rows],
        start_col + left,
        removed_title_hints,
    )


def _score_candidate(
    headers: list[str], rows: list[list[Any]], profiles: list[ColumnProfile], title: str
) -> int:
    score = 24.0
    score += min(20.0, math.log2(max(2, len(rows))) * 3.1)
    score += min(18.0, sum(1 for profile in profiles if profile.role in {"metric", "delta"}) * 4.0)
    score += min(10.0, sum(1 for header in headers if re.search(r"[A-Za-z]", header)) * 1.2)
    if title:
        score += 8.0
    if any(_contains_hint(header, METRIC_HINTS) for header in headers):
        score += 8.0
    if any(profile.data_type == "size" for profile in profiles):
        score += 5.0
    long_headers = sum(1 for header in headers if len(header) > 90)
    score -= long_headers * 3.0
    if len(rows) <= 2:
        score -= 12.0
    if len(headers) > 18:
        score -= min(12.0, (len(headers) - 18) * 0.7)
    return max(1, min(99, int(round(score))))


def _derive_suggestions(
    headers: list[str], rows: list[list[Any]], profiles: list[ColumnProfile]
) -> tuple[int, list[int], list[int], str, str, bool, str, str, list[str]]:
    dimension_candidates = [
        profile for profile in profiles if profile.role == "dimension" and profile.unique_count > 1
    ]
    if not dimension_candidates:
        dimension_candidates = [profile for profile in profiles if profile.role == "dimension"]
    if not dimension_candidates:
        dimension_candidates = [profile for profile in profiles if profile.role != "annotation"]
    category_profile = max(dimension_candidates, key=lambda profile: profile.dimension_score)
    category_index = category_profile.index

    metric_profiles = [profile for profile in profiles if profile.role == "metric" and profile.index != category_index]
    delta_profiles = [profile for profile in profiles if profile.role == "delta" and profile.index != category_index]
    metric_profiles.sort(key=lambda profile: (-profile.metric_score, profile.index))

    category_repeats = category_profile.unique_count < max(2, int(len(rows) * 0.85))
    long_form = category_repeats and len([profile for profile in profiles if profile.role == "dimension"]) >= 2
    metric_limit = 2 if long_form or len(profiles) > 12 or (len(rows) > 100 and len(metric_profiles) > 4) else 8
    suggested_metrics = metric_profiles[:metric_limit]
    # Avoid auto-charting unrelated units on one axis (for example MFU % with days,
    # or packet throughput with frame-loss counts). Generic series labels such as
    # Baseline/Candidate or 1QP/2QP remain grouped together.
    if len(suggested_metrics) > 1:
        primary_unit = _metric_unit_key(suggested_metrics[0].header)
        if primary_unit != "generic":
            same_unit = [
                profile for profile in metric_profiles
                if _metric_unit_key(profile.header) == primary_unit
            ]
            if same_unit:
                suggested_metrics = same_unit[:metric_limit]
    if len(suggested_metrics) > 1 and all(_metric_unit_key(profile.header) == "generic" for profile in suggested_metrics):
        suggested_metrics = sorted(suggested_metrics, key=lambda profile: profile.index)
    if not suggested_metrics:
        suggested_metrics = delta_profiles[:4]
    series_indexes = [profile.index for profile in suggested_metrics]

    group_candidates = [
        profile
        for profile in profiles
        if profile.role == "dimension"
        and profile.index != category_index
        and not profile.identifier
        and 2 <= profile.unique_count <= 10
    ]
    for profile in group_candidates:
        normalized = _normalized_header(profile.header)
        if any(hint in normalized for hint in GROUP_HINTS):
            profile.dimension_score += 3.0
    group_candidates.sort(key=lambda profile: (-profile.dimension_score, profile.unique_count, profile.index))
    group_by = [group_candidates[0].index] if long_form and group_candidates else []

    duplicate_key_count: dict[tuple[str, ...], int] = {}
    for row in rows:
        category = _clean_text(row[category_index]) if category_index < len(row) else ""
        group = tuple(_clean_text(row[index]) for index in group_by if index < len(row))
        key = (category, *group)
        duplicate_key_count[key] = duplicate_key_count.get(key, 0) + 1
    has_duplicates = any(count > 1 for count in duplicate_key_count.values())
    has_trial_dimension = any(profile.identifier for profile in profiles)
    aggregation = "median" if has_duplicates and has_trial_dimension else ("mean" if has_duplicates else "none")

    category_count = _category_unique_count(category_profile)
    category_header = headers[category_index] if category_index < len(headers) else ""
    category_header_norm = _normalized_header(category_header)
    ordered_category = category_profile.data_type in {"size", "number", "date"} or any(
        hint in category_header_norm for hint in ("size", "message", "input", "packet", "frame", "date", "time")
    )
    if ordered_category and category_count > 7:
        chart_type = "line"
    elif not ordered_category and category_count > 36:
        chart_type = "data_matrix"
    else:
        chart_type = "grouped_bar"
    if series_indexes and all(profiles[index].role == "delta" for index in series_indexes) and category_count <= 36:
        chart_type = "difference"

    show_data_table = category_count <= 8 and len(series_indexes) <= 3
    x_axis_title = headers[category_index]
    y_axis_title = _suggest_y_axis([headers[index] for index in series_indexes])

    warnings: list[str] = []
    if len(rows) > 300:
        warnings.append("This data has many rows. Use a filter or combine repeated test results before creating a chart.")
    if category_repeats:
        warnings.append("Some labels appear more than once. Choose how the chart should use those repeated results.")
    if len(metric_profiles) > 8:
        warnings.append("Many value columns were found. Select only the values needed for this slide.")
    if not ordered_category and category_count > 16:
        warnings.append("This data has many different labels. Use a filter, a grouped bar chart, or a data matrix instead of a line chart.")
    if len(group_candidates) > 0 and not group_by and category_repeats:
        warnings.append("This data includes several test settings. Compare by one setting or filter to a specific setup.")
    if not series_indexes:
        warnings.append("The app could not clearly identify a value column. Review the selected labels and values.")
    return (
        category_index,
        series_indexes,
        group_by,
        chart_type,
        aggregation,
        show_data_table,
        x_axis_title,
        y_axis_title,
        warnings,
    )



def _infer_header_from_values(column_index: int, values: list[Any]) -> str:
    text = " ".join(_clean_text(value).lower() for value in values if not _is_blank(value))
    if column_index == 0:
        return "Test"
    if "rdma" in text and any(unit in text for unit in ("gb/s", "gbps", "mb/s", "mbps")):
        return "RDMA bandwidth"
    if "nvl" in text and any(unit in text for unit in ("gb/s", "gbps", "mb/s", "mbps")):
        return "NVL bandwidth"
    if "nvl chunk" in text:
        return "NVL chunk"
    if "rdma chunk" in text:
        return "RDMA chunk"
    return f"Column {column_index + 1}"


def _headerless_candidates(
    sheet_name: str,
    sheet_state: str,
    normalized: list[list[Any]],
    sheet_number: int,
    existing: list[_Candidate],
) -> list[_Candidate]:
    results: list[_Candidate] = []
    runs: list[tuple[int, int]] = []
    run_start: int | None = None
    for row_index, row in enumerate(normalized):
        populated = sum(1 for value in row if not _is_blank(value))
        numeric = sum(1 for value in row if _is_number(value))
        qualifies = populated >= 2 and numeric >= 1
        if qualifies and run_start is None:
            run_start = row_index
        elif not qualifies and run_start is not None:
            if row_index - run_start >= 3:
                runs.append((run_start, row_index - 1))
            run_start = None
    if run_start is not None and len(normalized) - run_start >= 3:
        runs.append((run_start, len(normalized) - 1))

    for run_start, run_end in runs:
        if any(
            max(0, min(run_end, item.r1) - max(run_start, item.r0) + 1) / max(1, run_end - run_start + 1) >= 0.6
            for item in existing
        ):
            continue
        nonempty_columns = sorted({
            column
            for row_index in range(run_start, run_end + 1)
            for column, value in enumerate(normalized[row_index])
            if not _is_blank(value)
        })
        if len(nonempty_columns) < 2:
            continue
        start_col, end_col = nonempty_columns[0], nonempty_columns[-1]
        if end_col - start_col + 1 > 12:
            continue
        rows = [normalized[row_index][start_col : end_col + 1] for row_index in range(run_start, run_end + 1)]
        numeric_columns = sum(
            1
            for column in range(end_col - start_col + 1)
            if sum(_is_number(row[column]) for row in rows) >= max(2, int(len(rows) * 0.6))
        )
        if numeric_columns < 1:
            continue
        columns = [[row[index] for row in rows] for index in range(end_col - start_col + 1)]
        headers = [_infer_header_from_values(index, values) for index, values in enumerate(columns)]
        profiles = [_profile_column(index, header, columns[index]) for index, header in enumerate(headers)]
        if not any(profile.role in {"metric", "delta"} for profile in profiles):
            continue
        (
            category_index,
            series_indexes,
            group_by,
            chart_type,
            aggregation,
            show_data_table,
            x_axis_title,
            y_axis_title,
            warnings,
        ) = _derive_suggestions(headers, rows, profiles)
        title = _nearest_title(normalized, run_start, start_col, end_col) or sheet_name
        title = title[:150]
        kind = _kind_for(sheet_name, title, headers)
        if y_axis_title == "Value" and kind in {"RCCL", "IB"}:
            y_axis_title = "Bandwidth"
        confidence = min(69, 48 + len(rows) * 2)
        warnings.insert(0, "Column headings were inferred from the data. Check the selected labels and values before generating.")
        table_id = f"s{sheet_number}-r{run_start + 1}-c{start_col + 1}-{end_col + 1}-inferred"
        table = DetectedTable(
            table_id=table_id,
            sheet_name=sheet_name,
            sheet_state=sheet_state,
            kind=kind,
            title=title,
            display_name=f"{sheet_name} · {title}" if title != sheet_name else sheet_name,
            header_row=run_start + 1,
            data_start_row=run_start + 1,
            data_end_row=run_end + 1,
            start_col=start_col + 1,
            end_col=end_col + 1,
            headers=headers,
            rows=rows,
            column_profiles=profiles,
            suggested_series_indexes=series_indexes,
            suggested_category_index=category_index,
            suggested_group_by_indexes=group_by,
            suggested_chart_type=chart_type,
            suggested_aggregation=aggregation,
            suggested_show_data_table=show_data_table,
            suggested_x_axis_title=x_axis_title,
            suggested_y_axis_title=y_axis_title,
            confidence=confidence,
            warnings=warnings,
        )
        results.append(_Candidate(table=table, r0=run_start, r1=run_end, c0=start_col, c1=end_col))
    return results

def _detect_on_sheet(
    sheet_name: str, sheet_state: str, values: list[list[Any]], sheet_number: int
) -> list[DetectedTable]:
    if not values:
        return []
    max_columns = max((len(row) for row in values), default=0)
    normalized = [row + [None] * (max_columns - len(row)) for row in values]
    candidates: list[_Candidate] = []

    for header_index, row in enumerate(normalized):
        nonempty_columns = [index for index, value in enumerate(row) if not _is_blank(value)]
        for start_col, end_col in _split_segments(nonempty_columns, max_gap=1):
            width = end_col - start_col + 1
            if width < 2 or width > 60:
                continue
            raw_headers = row[start_col : end_col + 1]
            header_nonempty = sum(1 for value in raw_headers if not _is_blank(value))
            header_text = sum(1 for value in raw_headers if not _is_blank(value) and not _is_number(value))
            header_numeric = sum(1 for value in raw_headers if _is_number(value))
            header_errors = sum(
                1 for value in raw_headers if isinstance(value, str) and value.strip().startswith("#")
            )
            semantic_header = any(
                _contains_hint(_clean_text(value), CATEGORY_HINTS | DIMENSION_HINTS | METRIC_HINTS | DELTA_HINTS)
                for value in raw_headers
                if not _is_blank(value)
            )
            if header_nonempty < 2 or header_text < 1 or header_errors:
                continue
            if header_numeric / header_nonempty >= 0.5 and not semantic_header:
                continue
            if width <= 4 and header_numeric / header_nonempty >= 0.5:
                continue
            candidate_rows, row_numbers, rows_truncated = _collect_rows(normalized, header_index, start_col, end_col)
            if len(candidate_rows) < 2:
                continue
            raw_headers, candidate_rows, adjusted_start_col, removed_title_hints = _trim_empty_edges(
                raw_headers, candidate_rows, start_col
            )
            if len(raw_headers) < 2:
                continue
            adjusted_end_col = adjusted_start_col + len(raw_headers) - 1
            headers = _unique_headers(raw_headers)
            columns = [
                [data_row[index] if index < len(data_row) else None for data_row in candidate_rows]
                for index in range(len(headers))
            ]
            profiles = [_profile_column(index, header, columns[index]) for index, header in enumerate(headers)]
            if not any(profile.role in {"metric", "delta"} for profile in profiles):
                continue
            graphable_values = sum(
                1
                for profile in profiles
                if profile.role in {"metric", "delta"}
                for value in columns[profile.index]
                if _is_number(value)
            )
            if graphable_values < 4:
                continue

            nearest_title = _nearest_title(normalized, header_index, adjusted_start_col, adjusted_end_col)
            hint_title = next((hint for hint in removed_title_hints if _meaningful_title(hint)), "")
            title = hint_title or nearest_title
            (
                category_index,
                series_indexes,
                group_by,
                chart_type,
                aggregation,
                show_data_table,
                x_axis_title,
                y_axis_title,
                warnings,
            ) = _derive_suggestions(headers, candidate_rows, profiles)
            if rows_truncated:
                warnings.insert(0, f"This table exceeds {MAX_TABLE_ROWS:,} data rows; only the first {MAX_TABLE_ROWS:,} were loaded.")
            if not title or _normalized_header(title) == _normalized_header(headers[category_index]):
                title = sheet_name
            title = title[:150]
            kind = _kind_for(sheet_name, title, headers)
            if y_axis_title == "Value" and kind in {"RCCL", "IB"}:
                y_axis_title = "Bandwidth"
            confidence = _score_candidate(headers, candidate_rows, profiles, title)
            if confidence < 35:
                continue
            if confidence < 58:
                warnings.insert(0, "The app was not fully confident about this data section. Check the headings and previewed rows.")
            display_name = f"{sheet_name} · {title}" if title != sheet_name else sheet_name
            table_id = f"s{sheet_number}-r{header_index + 1}-c{adjusted_start_col + 1}-{adjusted_end_col + 1}"
            table = DetectedTable(
                table_id=table_id,
                sheet_name=sheet_name,
                sheet_state=sheet_state,
                kind=kind,
                title=title,
                display_name=display_name,
                header_row=header_index + 1,
                data_start_row=row_numbers[0] + 1,
                data_end_row=row_numbers[-1] + 1,
                start_col=adjusted_start_col + 1,
                end_col=adjusted_end_col + 1,
                headers=headers,
                rows=candidate_rows,
                column_profiles=profiles,
                suggested_series_indexes=series_indexes,
                suggested_category_index=category_index,
                suggested_group_by_indexes=group_by,
                suggested_chart_type=chart_type,
                suggested_aggregation=aggregation,
                suggested_show_data_table=show_data_table,
                suggested_x_axis_title=x_axis_title,
                suggested_y_axis_title=y_axis_title,
                confidence=confidence,
                warnings=warnings,
            )
            candidates.append(
                _Candidate(
                    table=table,
                    r0=header_index,
                    r1=row_numbers[-1],
                    c0=adjusted_start_col,
                    c1=adjusted_end_col,
                )
            )

    candidates.extend(_headerless_candidates(sheet_name, sheet_state, normalized, sheet_number, candidates))

    selected: list[_Candidate] = []
    for candidate in sorted(candidates, key=lambda item: (-item.table.confidence, item.r0, item.c0)):
        candidate_area = (candidate.r1 - candidate.r0 + 1) * (candidate.c1 - candidate.c0 + 1)
        duplicate = False
        for accepted in selected:
            row_overlap = max(0, min(candidate.r1, accepted.r1) - max(candidate.r0, accepted.r0) + 1)
            col_overlap = max(0, min(candidate.c1, accepted.c1) - max(candidate.c0, accepted.c0) + 1)
            intersection = row_overlap * col_overlap
            accepted_area = (accepted.r1 - accepted.r0 + 1) * (accepted.c1 - accepted.c0 + 1)
            if intersection / max(1, min(candidate_area, accepted_area)) >= 0.65:
                duplicate = True
                break
        if not duplicate:
            selected.append(candidate)
    return [item.table for item in sorted(selected, key=lambda item: (item.r0, item.c0))]


def inspect_workbook(raw: bytes) -> tuple[dict[str, Any], list[DetectedTable]]:
    if not raw:
        raise ValueError("The workbook is empty.")
    workbook = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    tables: list[DetectedTable] = []
    sheet_summaries: list[dict[str, Any]] = []
    for sheet_number, worksheet in enumerate(workbook.worksheets, start=1):
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        while rows and all(_is_blank(value) for value in rows[-1]):
            rows.pop()
        sheet_tables = _detect_on_sheet(worksheet.title, worksheet.sheet_state, rows, sheet_number)
        tables.extend(sheet_tables)
        sheet_summaries.append(
            {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "table_count": len(sheet_tables),
                "recommended_count": sum(1 for table in sheet_tables if table.recommended),
            }
        )
    workbook.close()

    kind_counts: dict[str, int] = {}
    for table in tables:
        kind_counts[table.kind] = kind_counts.get(table.kind, 0) + 1
    metadata = {
        "worksheet_count": len(sheet_summaries),
        "worksheet_names": [sheet["name"] for sheet in sheet_summaries],
        "worksheets": sheet_summaries,
        "detected_table_count": len(tables),
        "recommended_table_count": sum(1 for table in tables if table.recommended),
        "kinds": dict(sorted(kind_counts.items())),
    }
    return metadata, tables
